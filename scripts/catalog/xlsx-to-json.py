import json, sys
import openpyxl

src = sys.argv[1] if len(sys.argv) > 1 else "oasis-songs.xlsx"
out = sys.argv[2] if len(sys.argv) > 2 else "oasis-songs.json"

wb = openpyxl.load_workbook(src)
ws = wb.active

def na(v):
    if v is None: return None
    s = str(v).strip()
    return None if s == "" or s.lower() == "n/a" else s

rows = []
for r in range(2, ws.max_row + 1):
    title = na(ws.cell(r, 1).value)
    if not title:
        continue
    tags_raw = ws.cell(r, 6).value or ""
    rows.append({
        "title": title,
        "author": na(ws.cell(r, 2).value) or "",
        "key": na(ws.cell(r, 3).value) or "",
        "bpm": na(ws.cell(r, 4).value),
        "timeSig": na(ws.cell(r, 5).value) or "",
        "tags": [t.strip() for t in str(tags_raw).split(",") if t.strip()],
        "musicalUrl": na(ws.cell(r, 8).value),
        "lyricsUrl": na(ws.cell(r, 9).value),
    })

with open(out, "w", encoding="utf-8") as f:
    json.dump(rows, f, ensure_ascii=False, indent=2)
print(f"Wrote {len(rows)} rows to {out}")
